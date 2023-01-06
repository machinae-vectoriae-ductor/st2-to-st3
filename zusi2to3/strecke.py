#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import copy
import enum
import os
import math
import numpy as np
import xml.etree.ElementTree as ET
from collections import namedtuple
from openpyxl import Workbook

from . import landschaft
from .common import readfloat, readfloatstr


class RefTyp:
    AUFGLEISPUNKT = 0
    REGISTER = 2
    WEICHE = 3
    SIGNAL = 4
    AUFLOESEPUNKT = 5
    SIGNAL_GEGENRICHTUNG = 8
    WEICHE_GEGENRICHTUNG = 9


def get_ref_nr(elem_nr, ref_typ):
    return 10 * elem_nr + ref_typ


def allocate_refpunkt(nstrecke_strecke, elem_nr, ref_typ):
    nstrecke_re = ET.SubElement(nstrecke_strecke, "ReferenzElemente")
    nstrecke_re.attrib["ReferenzNr"] = str(get_ref_nr(elem_nr, ref_typ))
    nstrecke_re.attrib["StrElement"] = str(elem_nr)
    if ref_typ == RefTyp.SIGNAL_GEGENRICHTUNG:
        nstrecke_re.attrib["RefTyp"] = str(RefTyp.SIGNAL)
    elif ref_typ == RefTyp.WEICHE_GEGENRICHTUNG:
        nstrecke_re.attrib["RefTyp"] = str(RefTyp.WEICHE)
    else:
        nstrecke_re.attrib["RefTyp"] = str(ref_typ)
        nstrecke_re.attrib["StrNorm"] = "1"

    return nstrecke_re


VerknParameter = namedtuple(
    "VerknParameter", ["dateiname_zusi", "x", "y", "z", "rx", "ry", "rz", "boundingr"]
)

Ereignis = namedtuple("Ereignis", ["nummer", "text", "wert"])


class Koordinate:
    def __init__(self):
        self.x = 0.0
        self.y = 0.0
        self.z = 0.0
        self.rx = 0.0
        self.ry = 0.0
        self.rz = 0.0
        self.lsdatei = ""


class SignalKoordinate:
    def __init__(self):
        self.x = 0.0
        self.y = 0.0
        self.z = 0.0
        self.rx = 0.0
        self.ry = 0.0
        self.rz = 0.0
        self.lsdatei = ""


class Signal:
    def __init__(self):
        self.elnr = 0
        self.btrst = ""
        self.stw = ""
        self.signame = ""
        self.sigflag = 0
        self.sigtyp = 0
        self.boundingr = 0.0
        self.koord = Koordinate
        self.sigframe = []
        self.block = ""
        self.gleis = ""
        self.matrix = []
        self.vsig_geschw = []
        self.hsig_geschw = []
        self.vsigs = []
        self.master = 0


class MatrixZeile:
    def __init__(self):
        self.block = ""
        self.gleis = ""
        self.vmax = 0
        self.fstrtyp = 0
        self.spalten = []


class MatrixEintrag:
    def __init__(self):
        self.bild = 0
        self.vmax = 0
        self.id = 0
        self.er1 = 0
        self.er2 = 0


signalleer = {
    "Vorhanden": False,  # Signal in dem Streckenelement vorhanden
    "BoundingR": 0.0,  # Boundingradius
    "koord": SignalKoordinate,  # Standortkoordinate
    "ls0": "",  # Datei, die die zum Signalbild gehörende Landschaft enthält
    "ls1": "",  # Datei, die die zum Signalbild gehörende Landschaft enthält
    "Ereignis": Ereignis,  # Ereignis
    "EreignisWert": 0.0,  # Ereigniswert
    "vsig": 0.0,  # Am Signal angekündigte Geschwindigkeit
    "master": 0,  # Verweis auf Masterelement
}


def conv_ereignis(z2ernr, z2ertext):
    z3ernr = 0
    z3ertext = ""
    z3erwert = 0.0
    if z2ernr == 0:  # Kein Ereignis
        pass
    # Bedingte Entgleisung, wird ausgelöst bei "Fahrt-Geschwindigkeit in km/h größer Ereigniswert" (+Toleranz)
    elif z2ernr >= 1 and z2ernr <= 499:
        z3ernr = 1
        z3erwert = 1.1 * z2ernr / 3.6
    # PZB 500 Hz-Beeinflussung
    elif z2ernr == 500:
        z3ernr = 500
    # PZB 1000 Hz-Beeinflussung
    elif z2ernr == 1000:
        z3ernr = 1000
    # Bedingte 1000 Hz-PZB-Beeinflussung, wird ausgelöst bei "Fahrt-Geschwindigkeit in km/h größer (Ereigniswert - 1000)", also z.B. 1105: 1000 Hz Beeinflussung bei 105 km/h und mehr
    elif z2ernr >= 1001 and z2ernr <= 1500:
        z3ernr = 1000
        z3erwert = float(z2ernr - 1000)
    # PZB 2000 Hz-Beeinflussung
    elif z2ernr == 2000:
        z3ernr = 2000
    # Bedingte 2000 Hz-PZB-Beeinflussung (Geschwindigkeitsprüfabschnitt), wird ausgelöst bei "Fahrt-Geschwindigkeit in km/h größer (Ereigniswert - 2000)"
    elif z2ernr >= 2001 and z2ernr <= 2500:
        z3ernr = 2000
        z3erwert = float(z2ernr - 2000)
    # Fahrstraße anfordern (wird standardmäßig nicht gebraucht, da die Züge automatisch anfordern)
    elif z2ernr == 3001:
        z3ernr = 5
    # Fahrstraße auflösen
    elif z2ernr == 3002:
        z3ernr = 4
    # Zug entfernen (Zug wird sofort entfernt und der belegte Blockabschnitt freigegeben)
    elif z2ernr == 3003:
        z3ernr = 16
    # Zwangshalt
    elif z2ernr == 3004:
        z3ernr = 14
    # Langsamfahrt Ende
    elif z2ernr == 3005:
        # z3ernr = 1000002
        pass
    # Betriebsstelle
    elif z2ernr == 3006:
        z3ernr = 6
        z3ertext = z2ertext
    # Haltepunkt erwarten
    elif z2ernr == 3007:
        # print("Ereignis Haltepunkt erwarten")
        pass
    # Bahnsteigmitte
    elif z2ernr == 3008:
        z3ernr = 1000011
        z3ertext = z2ertext
    # Bahnsteigende
    elif z2ernr == 3009:
        z3ernr = 1000012
        z3ertext = z2ertext
    # Langsamfahrt Anfang
    elif z2ernr == 3010:
        pass
    # Pfeifen, Kurz vorher Pfeifen - Missachtung führt zu Punktabzug
    elif z2ernr == 3011:
        z3ernr = 30
    # LZB-Anfang (Zug wird in die LZB aufgenommen)
    elif z2ernr == 3012:
        z3ernr = 1003001
    # LZB-Ende (Zug wird am vorhergehenden Hauptsignal aus der LZB entlassen)
    elif z2ernr == 3013:
        z3ernr = 3002
    # Vorher keine Fahrstraße (Bis der Zug dieses Ereignis überquert hat, wird ihm keine Fahrstraße mehr gestellt)
    elif z2ernr == 3021:
        z3ernr = 13
    # Zp9-Signal (muss in den Eigenschaften eines Fahrstraßensignals gesetzt werden, dann wird dieses erst beim Abfahrtauftrag auf "Fahrt" gestellt)
    elif z2ernr == 3022:
        z3ernr = 25
    # Weiterfahrt nach Halt, Überfahren ohne vorigen Halt führt zu Punktabzug
    elif z2ernr == 3023:
        z3ernr = 18
    # Signum Warnung
    # elif z2ernr == 3024:
    # pass
    # Signum Halt
    # elif z2ernr == 3025:
    # pass
    # Sich nähernder Zug bekommt frühestens 1000m vor diesem Ereignis die nächste Fahrstraße
    # elif z2ernr == 3026:
    # pass
    # Sich nähernder Zug bekommt frühestens 2000m vor diesem Ereignis die nächste Fahrstraße
    # elif z2ernr == 3027:
    # pass
    # Sich nähernder Zug bekommt frühestens 3000m vor diesem Ereignis die nächste Fahrstraße
    # elif z2ernr == 3028:
    # pass
    # Vorher keine Vorsignalverknüpfung - Hauptsignal ignoriert beim Verknüpfen der Signallogik alle Vorsignale vor diesem Ereignis
    # elif z2ernr == 3029:
    # z3ernr = 20
    # pass
    # Ereignis ohne Funktion
    elif z2ernr == 3030:
        pass
    # Befehl A
    # elif z2ernr == 3031:
    # z3ernr = 32
    # z3erwert =
    # pass
    # Befehl A (Stillstand)
    # elif z2ernr == 3032:
    # z3ernr = 32
    # z3erwert =
    # pass
    # Befehl B
    # elif z2ernr == 3033:
    # z3ernr = 32
    # z3erwert =
    # pass
    # Befehl B (Stillstand)
    # elif z2ernr == 3034:
    # z3ernr = 32
    # z3erwert =
    # pass
    # Langsamfahrtende (Zuganfang)
    # elif z2ernr == 3035:
    # pass
    # Wendepunkt (wird nur im Fahrplaneditor benötigt, Kennzeichnet Wendepunkte, muß hinter dem Wende-HSig liegen)
    elif z2ernr == 3036:
        pass
    # Wendepunkt auf anderen Blocknamen (wenn das Hsig vor dem Wendemanöver einen anderen Blocknamen hat als das nach dem Wenden)
    elif z2ernr == 3037:
        pass
    # Signal ist zugbedient (muß in den Eigenschaften eines Fahrstraßensignals gesetzt werden, dann wird dieses durch Ereignis 3039 gestellt)
    # elif z2ernr == 3038:
    # pass
    # Zugbedientes Signal schalten
    # elif z2ernr == 3039:
    # pass
    # Streckensound, Sound-Datei unter "Beschreibung" angeben
    # elif z2ernr == 3040:
    # pass
    # Abrupt-Halt: Zug wird schlagartig gestoppt
    elif z2ernr == 3041:
        z3ernr = 15
    # GNT: Keine Geschwindigkeitserhöhung
    elif z2ernr == 4000:
        pass
    # GNT-Anfang
    elif z2ernr == 4001:
        z3ernr = 4011
        pass
    # GNT-Ende
    elif z2ernr == 4002:
        z3ernr = 4012
        pass
    # GNT: PZB-Unterdrückung auf 150 m (alle PZB-Magnete werden auf den nächsten 150 m unterdrückt)
    elif z2ernr == 4003:
        z3ernr = 4013
        pass
    # GNT: Erhöhung der GNT-Geschwindigkeit gegenüber normaler Geschwindigkeit
    # elif z2ernr >= 4004 and z2ernr <= 4500:
    # z3ernr = 4010
    # z3erwert =
    # pass
    # elif z2ernr > 0:
    #     print("Ereignis: ", z2ernr, z2ertext)
    # Rechtsbündiger Eintrag
    if "..." in z3ertext:
        z3ertext = z3ertext.replace("...", "@sep=") + "@"
    # Verkürzter Vorsignalabstand
    if chr(186) in z3ertext:
        z3ertext = z3ertext.replace(chr(186), "@icon=7@")
    return Ereignis(z3ernr, z3ertext, z3erwert)


def define_circle(p1, p2, p3):
    """
    Returns the center and radius of the circle passing the given 3 points.
    In case the 3 points form a line, returns (None, infinity).
    """
    temp = p2[0] * p2[0] + p2[1] * p2[1]
    bc = (p1[0] * p1[0] + p1[1] * p1[1] - temp) / 2
    cd = (temp - p3[0] * p3[0] - p3[1] * p3[1]) / 2
    det = (p1[0] - p2[0]) * (p2[1] - p3[1]) - (p2[0] - p3[0]) * (p1[1] - p2[1])

    if abs(det) < 1.0e-6:
        return (None, np.inf)

    # Center of circle
    cx = (bc * (p2[1] - p3[1]) - cd * (p1[1] - p2[1])) / det
    cy = ((p1[0] - p2[0]) * cd - (p2[0] - p3[0]) * bc) / det

    radius = np.sqrt((cx - p1[0]) ** 2 + (cy - p1[1]) ** 2)
    return ((cx, cy), radius)


def conv_str(z2pfad, z2streckendateiname_rel, z3pfad, zielverzeichnis_rel):
    streckenelemente = {}
    nodes = {}
    signale = {}
    anonymesignale = {}
    fahrstrsignale = set()
    regnr = 20000  # TODO

    streckenname = os.path.splitext(os.path.basename(z2streckendateiname_rel))[0]
    z2streckendateiname_abs = os.path.join(z2pfad, "Strecken", z2streckendateiname_rel)
    z3streckendateiname_rel = os.path.join(
        "Routes", zielverzeichnis_rel, streckenname + ".st3"
    )
    z3streckendateiname_abs = os.path.join(z3pfad, z3streckendateiname_rel)

    nstrecke_root = ET.Element("Zusi")
    tree = ET.ElementTree(nstrecke_root)
    nstrecke_info = ET.SubElement(
        nstrecke_root,
        "Info",
        {"DateiTyp": "Strecke", "Version": "A.4", "MinVersion": "A.1"},
    )
    nstrecke_strecke = ET.SubElement(nstrecke_root, "Strecke", {"Himmelsmodell": "2"})

    f = open(z2streckendateiname_abs, "r", encoding="iso-8859-1")
    zusiversion = f.readline().strip()
    if zusiversion != "2.3":
        print("Version", zusiversion, "wird nicht gelesen")
        sys.exit()

    strecke_autor = f.readline().strip()
    if strecke_autor:
        ET.SubElement(nstrecke_info, "AutorEintrag", {"AutorName": strecke_autor})
    print("Autor:", strecke_autor)
    print("Breitengrad:", f.readline())  # Breitengrad

    rekursionstiefe = int(f.readline())

    while not (strecke_info := f.readline().strip()).startswith(
        "#"
    ):  # Freier Text, nur zur Info
        print("Info:", strecke_info)
    while not (strecke_utm := f.readline().strip()).startswith(
        "#"
    ):  # UTM-Infos, werden von den Ziegler Tools abgelegt
        print("UTM:", strecke_utm)

    f.readline()  # Geschwindigkeits-Multiplikator bezogen auf m/s (bisher ohne Funktion)
    ls_datei = landschaft.conv_ls(f.readline().strip(), no_displacement=True)[
        0
    ]  # Landschaftsdatei der Strecke
    print("Landschaftsdatei:", ls_datei)

    ET.SubElement(nstrecke_strecke, "Datei", {"Dateiname": ls_datei})
    nstrecke_skydome = ET.SubElement(nstrecke_strecke, "SkyDome")
    ET.SubElement(nstrecke_skydome, "HimmelTex", {"Dateiname": "_Setup\sky\sky.dds"})
    ET.SubElement(nstrecke_skydome, "SonneTex", {"Dateiname": "_Setup\sky\sun.dds"})
    ET.SubElement(
        nstrecke_skydome,
        "SonneHorizontTex",
        {"Dateiname": "_Setup\sky\sun_horizon.dds"},
    )
    ET.SubElement(nstrecke_skydome, "MondTex", {"Dateiname": "_Setup\sky\moon.dds"})
    ET.SubElement(nstrecke_skydome, "SternTex", {"Dateiname": "_Setup\sky\star.dds"})

    print("Aufgleispunkte eintragen")
    aufgleispunkte = {}
    while not (refnr := f.readline()).startswith("#"):
        elem_nr = int(f.readline())
        aufgleispunkte[int(refnr)] = elem_nr
        beschr = f.readline().strip()
        nstrecke_re = allocate_refpunkt(nstrecke_strecke, elem_nr, RefTyp.AUFGLEISPUNKT)
        # print(elem_nr, beschr, RefTyp.AUFGLEISPUNKT)
        nstrecke_re.attrib["Info"] = beschr

    print("Streckenstandorte eintragen")
    while (str_stdort_y := readfloatstr(f)) :
        str_stdort_x = readfloatstr(f)  # Streckenstandort y-Koordinate
        str_stdort_z = readfloatstr(f)  # Streckenstandort z-Koordinate
        str_lookat_y = str(
            float(str_stdort_y) - 30 * math.sin(readfloat(f))
        )  # Streckenstandort Drehung um y-Koordinate in rad
        str_lookat_x = str(
            float(str_stdort_x) - 30 * math.sin(readfloat(f))
        )  # Streckenstandort Drehung um x-Koordinate in rad
        str_lookat_z = str(
            float(str_stdort_z) - 30 * math.sin(readfloat(f))
        )  # Streckenstandort Drehung um z-Koordinate in rad
        # str_up_x = str(float(str_stdort_x) + 2.0)
        # str_up_y = str(float(str_stdort_y) + 2.0)
        # str_up_z = str(float(str_stdort_z) + 2.0)
        str_stdort_name = f.readline().strip()
        # print("Streckenstandort:", str_stdort_name)
        # nstrecke_stdort = ET.SubElement(nstrecke_strecke, "StreckenStandort", {"StrInfo": str_stdort_name})
        # ET.SubElement(nstrecke_stdort, "p", {"X": str_stdort_x, "Y": str_stdort_y, "Z": str_stdort_z})
        # ET.SubElement(nstrecke_stdort, "lookat", {"X": str_lookat_x, "Y": str_lookat_y, "Z": str_lookat_z})
        # ET.SubElement(nstrecke_stdort, "up", {"X": str_up_x, "Y": str_up_y, "Z": str_up_z})

    while True:
        hilf_string = f.readline()
        if hilf_string == "":
            break
        else:
            elementnr = int(hilf_string)
        streckenelemente[elementnr] = {"Anschluss": 0xFF00}
        streckenelemente[elementnr]["Normkm"] = round(readfloat(f) * 0.001, 6)
        if f.readline().strip() == "+":
            streckenelemente[elementnr]["Normpos"] = 1
        # Landschaftsbezeichnung
        landschaftsbezeichnung = f.readline().strip()
        # if len(landschaftsbezeichnung) > 1:
        #     print(elementnr, landschaftsbezeichnung)
        # Ereignis
        ereignis = int(f.readline())
        # x-Anfangs-Standortkoordinate
        hilf_float = readfloat(f)
        if hilf_float != 0.0:
            streckenelemente[elementnr]["gx"] = hilf_float
        # y-Anfangs-Standortkoordinate
        hilf_float = readfloat(f)
        if hilf_float != 0.0:
            streckenelemente[elementnr]["gy"] = hilf_float
        # z-Anfangs-Standortkoordinate
        hilf_float = readfloat(f)
        if hilf_float != 0.0:
            streckenelemente[elementnr]["gz"] = hilf_float
        # x-End-Standortkoordinate
        hilf_float = readfloat(f)
        if hilf_float != 0.0:
            streckenelemente[elementnr]["bx"] = hilf_float
        # y-End-Standortkoordinate
        hilf_float = readfloat(f)
        if hilf_float != 0.0:
            streckenelemente[elementnr]["by"] = hilf_float
        # z-End-Standortkoordinate
        hilf_float = readfloat(f)
        if hilf_float != 0.0:
            streckenelemente[elementnr]["bz"] = hilf_float
        # Überhöhung in rad
        hilf_float = readfloat(f)
        if hilf_float != 0.0:
            streckenelemente[elementnr]["Ueberh"] = hilf_float
        # Nummern der nachfolgenden Streckenelemente
        succ = [
            x
            for x in [int(f.readline()), int(f.readline()), int(f.readline())]
            if x != 0
        ]
        streckenelemente[elementnr]["NachNorm"] = succ
        # Referenzpunkt für Weiche eintragen
        if len(succ) > 1:
            allocate_refpunkt(nstrecke_strecke, elementnr, RefTyp.WEICHE)
            # print(elementnr, get_ref_nr(elementnr, RefTyp.WEICHE), RefTyp.WEICHE)
        # Zulässige Geschwindigkeit in km/h
        hilf_float = readfloat(f)
        if hilf_float != 0.0:
            streckenelemente[elementnr]["NormvMax"] = hilf_float / 3.6
        f.readline()  # reserviert für spätere Funktionen
        # Name des Bahnsteigs oder der Betriebsstelle
        betriebsstelle = f.readline().strip()
        hilf_ereignis = conv_ereignis(ereignis, betriebsstelle)
        if hilf_ereignis.nummer > 0:
            NormEreignis = {}
            NormEreignis["Er"] = hilf_ereignis.nummer
            if len(hilf_ereignis.text) > 0:
                NormEreignis["Beschr"] = hilf_ereignis.text
            if hilf_ereignis.wert > 0.0:
                NormEreignis["Wert"] = hilf_ereignis.wert
            if "NormEreignis" in streckenelemente[elementnr]:
                streckenelemente[elementnr]["NormEreignis"].append(NormEreignis)
            else:
                streckenelemente[elementnr]["NormEreignis"] = [NormEreignis]
        if hilf_ereignis.nummer != 6 and len(hilf_ereignis.text) > 0:
            NormEreignis = {}
            NormEreignis["Er"] = 6
            NormEreignis["Beschr"] = hilf_ereignis.text
            if "NormEreignis" in streckenelemente[elementnr]:
                streckenelemente[elementnr]["NormEreignis"].append(NormEreignis)
            else:
                streckenelemente[elementnr]["NormEreignis"] = [NormEreignis]
        lage = f.readline().strip()  # Gibt an, ob Element eine besondere Lage hat
        if "T" in lage:
            streckenelemente[elementnr]["fkt"] = 1
        # Spannung Oberleitung in kV (0: nicht elektrifiziert)
        uol = f.readline().strip()
        if uol == "15":
            streckenelemente[elementnr]["Volt"] = 2
        elif uol == "25":
            streckenelemente[elementnr]["Volt"] = 3

        # Ab hier Eintragen in die XML-Datei

        nstrecke_str_element = ET.SubElement(nstrecke_strecke, "StrElement")
        nodes[elementnr] = nstrecke_str_element
        nstrecke_str_element.attrib["Nr"] = str(elementnr)
        if "Ueberh" in streckenelemente[elementnr]:
            nstrecke_str_element.attrib["Ueberh"] = str(
                streckenelemente[elementnr]["Ueberh"]
            )
        if "NormvMax" in streckenelemente[elementnr]:
            nstrecke_str_element.attrib["spTrass"] = str(streckenelemente[elementnr]["NormvMax"])
        else:
            nstrecke_str_element.attrib["spTrass"] = "150.0"
        if "Anschluss" in streckenelemente[elementnr]:
            nstrecke_str_element.attrib["Anschluss"] = str(
                streckenelemente[elementnr]["Anschluss"]
            )
        nstrecke_str_element.attrib["Oberbau"] = "Zusi 2"
        nstrecke_g = ET.SubElement(nstrecke_str_element, "g")
        if "gx" in streckenelemente[elementnr]:
            nstrecke_g.attrib["X"] = str(streckenelemente[elementnr]["gx"])
        if "gy" in streckenelemente[elementnr]:
            nstrecke_g.attrib["Y"] = str(streckenelemente[elementnr]["gy"])
        if "gz" in streckenelemente[elementnr]:
            nstrecke_g.attrib["Z"] = str(streckenelemente[elementnr]["gz"])
        nstrecke_b = ET.SubElement(nstrecke_str_element, "b")
        if "bx" in streckenelemente[elementnr]:
            nstrecke_b.attrib["X"] = str(streckenelemente[elementnr]["bx"])
        if "by" in streckenelemente[elementnr]:
            nstrecke_b.attrib["Y"] = str(streckenelemente[elementnr]["by"])
        if "bz" in streckenelemente[elementnr]:
            nstrecke_b.attrib["Z"] = str(streckenelemente[elementnr]["bz"])
        nstrecke_norm = ET.SubElement(nstrecke_str_element, "InfoNormRichtung")
        if "NormEreignis" in streckenelemente[elementnr]:
            for i in range(len(streckenelemente[elementnr]["NormEreignis"])):
                nstrecke_norm_er = ET.SubElement(
                    nstrecke_norm,
                    "Ereignis",
                    {"Er": str(streckenelemente[elementnr]["NormEreignis"][i]["Er"])},
                )
                if "Beschr" in streckenelemente[elementnr]["NormEreignis"][i]:
                    nstrecke_norm_er.attrib["Beschr"] = str(
                        streckenelemente[elementnr]["NormEreignis"][i]["Beschr"]
                    )
                if "Wert" in streckenelemente[elementnr]["NormEreignis"][i]:
                    nstrecke_norm_er.attrib["Wert"] = str(
                        streckenelemente[elementnr]["NormEreignis"][i]["Wert"]
                    )
        if "Normkm" in streckenelemente[elementnr]:
            nstrecke_norm.attrib["km"] = str(streckenelemente[elementnr]["Normkm"])
        if "Normpos" in streckenelemente[elementnr]:
            nstrecke_norm.attrib["pos"] = str(streckenelemente[elementnr]["Normpos"])

        if "NormvMax" in streckenelemente[elementnr]:
            nstrecke_norm.attrib["vMax"] = str(streckenelemente[elementnr]["NormvMax"])
        if "Volt" in streckenelemente[elementnr]:
            nstrecke_str_element.attrib["Volt"] = str(
                streckenelemente[elementnr]["Volt"]
            )
            if (
                streckenelemente[elementnr]["Volt"] == 2
                or streckenelemente[elementnr]["Volt"] == 3
            ):
                nstrecke_str_element.attrib["Drahthoehe"] = "5.4500"

        # Ab hier Einlesen der Signalinformationen

        signal = signalleer.copy()
        if (string := f.readline().strip()) != "#":
            signal["Vorhanden"] = True
            signal["koord"].x = float(string.replace(",", "."))
            signal["koord"].y = readfloat(f)
            signal["koord"].z = readfloat(f)
            signal["koord"].rx = readfloat(f)
            signal["koord"].ry = readfloat(f)
            signal["koord"].rz = readfloat(f)
            for i in range(6):
                f.readline()  # ohne Funktion
            sigframe_statisch = f.readline().strip()
            conv = landschaft.conv_ls(sigframe_statisch, no_displacement=True)
            signal["koord"].lsdatei = conv.dateiname_zusi
            signal["BoundingR"] = max(signal["BoundingR"], conv.boundingr)
            f.readline()  # ohne Funktion
            if not (sigframe_nicht_gestellt := f.readline()).startswith("#"):
                conv = landschaft.conv_ls(
                    sigframe_nicht_gestellt.strip(), no_displacement=True
                )
                signal["ls0"] = conv.dateiname_zusi
                signal["BoundingR"] = max(signal["BoundingR"], conv.boundingr)
                f.readline()  # ohne Funktion
                conv = landschaft.conv_ls(f.readline().strip(), no_displacement=True)
                signal["ls1"] = conv.dateiname_zusi
                signal["BoundingR"] = max(signal["BoundingR"], conv.boundingr)
                f.readline()  # ohne Funktion
                f.readline()  # Signalbilder-Endmarke
            signal["Ereignis"] = conv_ereignis(int(f.readline()), betriebsstelle)
            signal["vsig"] = int(f.readline())  # Am Signal angekündigte Geschwindigkeit
            signal["master"] = int(f.readline())
            # Solange Zusi 3 keine Funktion für stationäre Zp9-Signale hat, werden die Zusi 2-Zp9-Signale gelöscht.
            if signal["Ereignis"].nummer == 25:
                signal = signalleer.copy()

        # Ab hier Einlesen der Kombisignale

        Kombisignalvorhanden = False
        if (x1 := readfloat(f)) is not None:  # x-Standortkoordinate 1
            Kombisignalvorhanden = True
            sig = Signal()
            sig.elnr = elementnr

            y1 = readfloat(f)  # y-Standortkoordinate 1
            z1 = readfloat(f)  # z-Standortkoordinate 1
            rx1 = readfloatstr(f)  # Drehung am Standort um x-Achse 1
            ry1 = readfloatstr(f)  # Drehung am Standort um y-Achse 1
            rz1 = readfloatstr(f)  # Drehung am Standort um z-Achse 1

            x2 = readfloat(f)  # x-Standortkoordinate 2
            y2 = readfloat(f)  # y-Standortkoordinate 2
            z2 = readfloat(f)  # z-Standortkoordinate 2
            rx2 = readfloatstr(f)  # Drehung am Standort um x-Achse 2
            ry2 = readfloatstr(f)  # Drehung am Standort um y-Achse 2
            rz2 = readfloatstr(f)  # Drehung am Standort um z-Achse 2

            if not x1 and not y1 and not z1:
                sig.koord.x, sig.koord.y, sig.koord.z = x2, y2, z2
            elif not x2 and not y2 and not z2:
                sig.koord.x, sig.koord.y, sig.koord.z = x1, y1, z1
            else:
                sig.koord.x = (x1 + x2) / 2.0
                sig.koord.y = (y1 + y2) / 2.0
                sig.koord.z = (z1 + z2) / 2.0

            sig.koord.rx = 0.0
            sig.koord.ry = 0.0
            sig.koord.rz = 0.0

            # Erste .ls-Datei
            sigframes = []
            while not (lsdatei := f.readline().strip()).startswith("#"):
                sigframekoordinate = Koordinate()
                conv = landschaft.conv_ls(lsdatei, no_displacement=True)
                sigframekoordinate.lsdatei = conv.dateiname_zusi
                # Position
                if f.readline().startswith("2"):
                    sigframekoordinate.x = x2 - sig.koord.x
                    sigframekoordinate.y = y2 - sig.koord.y
                    sigframekoordinate.z = z2 - sig.koord.z
                    sigframekoordinate.rx = rx2
                    sigframekoordinate.ry = ry2
                    sigframekoordinate.rz = rz2
                else:
                    sigframekoordinate.x = x1 - sig.koord.x
                    sigframekoordinate.y = y1 - sig.koord.y
                    sigframekoordinate.z = z1 - sig.koord.z
                    sigframekoordinate.rx = rx1
                    sigframekoordinate.ry = ry1
                    sigframekoordinate.rz = rz1
                sig.boundingr = max(sig.boundingr, (conv.boundingr + (sigframekoordinate.x ** 2 + sigframekoordinate.y ** 2 + sigframekoordinate.z ** 2) ** 0.5))
                sig.sigframe.append(sigframekoordinate)
            # Hinzufügen der Bilder des Signales
            if signal["Vorhanden"]:
                sig.boundingr = max(sig.boundingr, signal["BoundingR"])
                # Signalframes eintragen
                sigframekoordinate = Koordinate()
                sigframekoordinate.lsdatei = signal["koord"].lsdatei
                sigframekoordinate.x = signal["koord"].x - sig.koord.x
                sigframekoordinate.y = signal["koord"].y - sig.koord.y
                sigframekoordinate.z = signal["koord"].z - sig.koord.z
                sigframekoordinate.rx = signal["koord"].rx
                sigframekoordinate.ry = signal["koord"].ry
                sigframekoordinate.rz = signal["koord"].rz
                sig.boundingr = max(sig.boundingr, (signal["BoundingR"] + (sigframekoordinate.x ** 2 + sigframekoordinate.y ** 2 + sigframekoordinate.z ** 2) ** 0.5))
                sig.sigframe.append(sigframekoordinate)
                positionls = len(sig.sigframe) - 1
                if len(signal["ls0"]) > 0:
                    sigframekoordinate = Koordinate()
                    sigframekoordinate.lsdatei = signal["ls0"]
                    sigframekoordinate.x = signal["koord"].x - sig.koord.x
                    sigframekoordinate.y = signal["koord"].y - sig.koord.y
                    sigframekoordinate.z = signal["koord"].z - sig.koord.z
                    sigframekoordinate.rx = signal["koord"].rx
                    sigframekoordinate.ry = signal["koord"].ry
                    sigframekoordinate.rz = signal["koord"].rz
                    sig.sigframe.append(sigframekoordinate)
                    sigframekoordinate = Koordinate()
                    sigframekoordinate.lsdatei = signal["ls1"]
                    sigframekoordinate.x = signal["koord"].x - sig.koord.x
                    sigframekoordinate.y = signal["koord"].y - sig.koord.y
                    sigframekoordinate.z = signal["koord"].z - sig.koord.z
                    sigframekoordinate.rx = signal["koord"].rx
                    sigframekoordinate.ry = signal["koord"].ry
                    sigframekoordinate.rz = signal["koord"].rz
                    sig.sigframe.append(sigframekoordinate)

            # Blockname
            sig.block = f.readline().strip()
            # Gleis
            sig.gleis = f.readline().strip()
            # Anzahl der Hauptsignalzeilen -1
            numzeilen = int(f.readline()) + 1
            # Anzahl der Vorsignalspalten -1
            numspalten = int(f.readline()) + 1

            sig.matrix = []

            # Hauptsignalgeschwindigkeiten
            seen_blocks = set()
            vmax0 = False
            vmaxm1 = False
            for i in range(0, numzeilen):
                # Fahrziel-Block, Fahrziel-Gleis, vmax, #, #
                mz = MatrixZeile()
                # Blockname des Fahrziels
                mz.block = f.readline().strip()
                # Gleis des Fahrziels
                mz.gleis = f.readline().strip()
                if mz.block or mz.gleis:
                    assert f"{mz.block} {mz.gleis}" not in seen_blocks
                    seen_blocks.add(f"{mz.block} {mz.gleis}")
                # zugeordnete Geschwindigkeit in km/h
                mz.vmax = int(f.readline())
                if mz.vmax == 0:
                    vmax0 = True
                # if mz.vmax == -1:
                #     vmaxm1 = True
                mz.fstrtyp = 6
                sig.matrix.append(mz)
                f.readline()
                f.readline()
            # Vorsignalgeschwindigkeiten in km/h
            for i in range(0, numspalten):
                vsig_geschw = int(f.readline())
                sig.vsig_geschw.append(vsig_geschw)

            # Aus bei Hp0
            f.readline()

            # Matrixeinträge

            for i in range(0, numzeilen):
                mz = sig.matrix[i]
                for j in range(0, numspalten):
                    me = MatrixEintrag()
                    # Bildauswahl, bitweise codiert (64 bit Wert)
                    me.bild = int(f.readline())
                    if signal["Vorhanden"]:
                        me.bild += 2 ** positionls
                        if len(signal["ls0"]) and sig.matrix[i].vmax == 0:
                            me.bild += 2 ** (positionls + 1)
                        else:
                            me.bild += 2 ** (positionls + 2)
                    ## Höchstgeschwindigkeit
                    me.vmax = int(f.readline())
                    # if me.vmax == 0 and sig.matrix[i].vmax != 0:
                    #     print(
                    #         f"Element {streckenelement['Nr']}, Zeile {i}, Spalte {j}: v=0, aber Zeile v!=0",
                    #         file=sys.stderr,
                    #     )
                    # ID
                    me.id = int(f.readline())
                    # Ereignis 1
                    me.er1 = conv_ereignis(int(f.readline()), betriebsstelle)
                    # if me.er1.nummer > 0:
                    #     print(elementnr, "1", me.er1)
                    # Ereignis 2
                    me.er2 = conv_ereignis(int(f.readline()), betriebsstelle)
                    # if me.er2.nummer > 0:
                    #     print(elementnr, "2", me.er2)
                    f.readline()

                    sig.matrix[i].spalten.append(me)

            if len(signal["ls0"]) and not vmax0:
                mz = MatrixZeile()
                mz.vmax = 0
                mz.fstrtyp = 1
                sig.matrix.append(mz)
                for j in range(0, numspalten):
                    me = MatrixEintrag()
                    me.bild = (
                        sig.matrix[0].spalten[j].bild
                        + 2 ** (positionls + 1)
                        - 2 ** (positionls + 2)
                    )
                    me.vmax = sig.matrix[0].spalten[j].vmax
                    me.id = sig.matrix[0].spalten[j].id
                    me.er1 = sig.matrix[0].spalten[j].er1
                    me.er2 = sig.matrix[0].spalten[j].er2
                    sig.matrix[i].spalten.append(me)

            # Ersatzsignale

            ersatz_bild = int(f.readline())
            ersatz_vmax = int(f.readline())
            ersatz_id = int(f.readline())
            ersatz_er1 = int(f.readline())
            ersatz_er2 = int(f.readline())
            ersatz_reserviert = f.readline()
            # Wahrscheinlichkeit Ersatzsignal
            f.readline()

            # Vorsignalstandorte

            vsig = f.readline()
            while not vsig.startswith("#"):
                sig.vsigs.append(int(vsig))
                vsig = f.readline()

            f.readline()  # reserviert

        # Eintragen des Signals als Kombisignal, wenn kein Kombisignal vorhanden ist.

        if Kombisignalvorhanden == False and signal["Vorhanden"]:
            fahrstrsignale.add(elementnr)
            Kombisignalvorhanden = True
            signal["Vorhanden"] = False
            sig = Signal()
            sig.elnr = elementnr
            sig.koord.x = signal["koord"].x
            sig.koord.y = signal["koord"].y
            sig.koord.z = signal["koord"].z
            sig.koord.rx = signal["koord"].rx
            sig.koord.ry = signal["koord"].ry
            sig.koord.rz = signal["koord"].rz
            sig.boundingr = signal["BoundingR"]
            # Signalframes eintragen
            sigframes = []
            sigframekoordinate = Koordinate()
            sigframekoordinate.lsdatei = signal["koord"].lsdatei
            sig.sigframe.append(sigframekoordinate)
            if len(signal["ls0"]) > 0:
                sigframekoordinate = Koordinate()
                sigframekoordinate.lsdatei = signal["ls0"]
                sig.sigframe.append(sigframekoordinate)
                sigframekoordinate = Koordinate()
                sigframekoordinate.lsdatei = signal["ls1"]
                sig.sigframe.append(sigframekoordinate)
                sig.sigflag = 4
                sig.sigtyp = 5
            else:
                sig.sigflag = 8

            # Matrixeinträge
            if len(signal["ls0"]) > 0:
                numzeilen = 2
            else:
                numzeilen = 1
            numspalten = 1
            # Hauptsignalgeschwindigkeiten
            for i in range(0, numzeilen):
                # Fahrziel-Block, Fahrziel-Gleis, vmax, #, #
                mz = MatrixZeile()
                # zugeordnete Geschwindigkeit in km/h
                if numzeilen == 2 and i == 0:
                    mz.vmax = 0
                else:
                    mz.vmax = -1
                if len(signal["ls0"]) > 0:
                    mz.fstrtyp = 2
                else:
                    mz.fstrtyp = 1
                sig.matrix.append(mz)
            # Vorsignalgeschwindigkeiten in km/h
            for i in range(0, numspalten):
                vsig_geschw = -1
                sig.vsig_geschw.append(vsig_geschw)

            # Matrixeinträge
            for i in range(0, numzeilen):
                mz = sig.matrix[i]
                for j in range(0, numspalten):
                    me = MatrixEintrag()
                    # Bildauswahl, bitweise codiert (64 bit Wert)
                    if numzeilen == 1:
                        me.bild = 1
                    elif i == 0:
                        me.bild = 3
                    else:
                        me.bild = 5
                    # Höchstgeschwindigkeit
                    me.vmax = -1
                    # Ereignis 1
                    me.er1 = signal["Ereignis"]
                    # Ereignis 2
                    me.er2 = conv_ereignis(0, betriebsstelle)

                    sig.matrix[i].spalten.append(me)

            # sig.master = signal["master"]

        # Kombissignale in XML-Datei eintragen

        if Kombisignalvorhanden:
            nstrecke_signal = ET.SubElement(nstrecke_norm, "Signal")
            if sig.block != "" and sig.gleis != "":
                nstrecke_signal.attrib["NameBetriebsstelle"] = sig.block
                nstrecke_signal.attrib["Stellwerk"] = sig.block
                nstrecke_signal.attrib["Signalname"] = sig.gleis
                signale[elementnr] = sig
            else:
                nstrecke_signal.attrib["Signalname"] = f"Element {elementnr}"
                anonymesignale[elementnr] = sig

            if sig.koord.x != 0.0:
                nstrecke_signal.attrib["SignalFlags"] = str(sig.sigflag)
            if sig.koord.x != 0.0:
                nstrecke_signal.attrib["SignalTyp"] = str(sig.sigtyp)
            nstrecke_signal.attrib["BoundingR"] = str(int(math.ceil(sig.boundingr)))
            nstrecke_p = ET.SubElement(nstrecke_signal, "p")
            if sig.koord.x != 0.0:
                nstrecke_p.attrib["X"] = str(sig.koord.x)
            if sig.koord.y != 0.0:
                nstrecke_p.attrib["Y"] = str(sig.koord.y)
            if sig.koord.z != 0.0:
                nstrecke_p.attrib["Z"] = str(sig.koord.z)
            nstrecke_phi = ET.SubElement(nstrecke_signal, "phi")
            if sig.koord.rx != 0.0:
                nstrecke_phi.attrib["X"] = str(sig.koord.rx)
            if sig.koord.ry != 0.0:
                nstrecke_phi.attrib["Y"] = str(sig.koord.ry)  # TODO warum?
            if sig.koord.rz != 0.0:
                nstrecke_phi.attrib["Z"] = str(sig.koord.rz)

            for i in range(len(sig.sigframe)):
                nstrecke_signalframe = ET.SubElement(nstrecke_signal, "SignalFrame")
                sigframes.append(nstrecke_signalframe)
                ET.SubElement(
                    nstrecke_signalframe,
                    "p",
                    {
                        "X": str(sig.sigframe[i].x),
                        "Y": str(sig.sigframe[i].y),
                        "Z": str(sig.sigframe[i].z),
                    },
                )
                ET.SubElement(
                    nstrecke_signalframe,
                    "phi",
                    {
                        "X": str(sig.sigframe[i].rx),
                        "Y": str(sig.sigframe[i].ry),
                        "Z": str(sig.sigframe[i].rz),
                    },
                )
                ET.SubElement(
                    nstrecke_signalframe,
                    "Datei",
                    {"Dateiname": sig.sigframe[i].lsdatei},
                )
            for i in range(0, len(sig.matrix)):
                if True:  # mz.vmax == 0 or mz.block or mz.gleis:
                    ET.SubElement(
                        nstrecke_signal,
                        "HsigBegriff",
                        {
                            "FahrstrTyp": str(mz.fstrtyp),
                            "HsigGeschw": "-1"
                            if sig.matrix[i].vmax == -1
                            else str(sig.matrix[i].vmax / 3.6),
                        },
                    )

            # if any(mz.vmax == 0 for mz in sig.matrix):
            #    ET.SubElement(nstrecke_norm, "Ereignis", {"Er":"29", "Beschr": f"{sig.block} {sig.gleis}"})

            for i in range(0, len(sig.vsig_geschw)):
                ET.SubElement(
                    nstrecke_signal,
                    "VsigBegriff",
                    {
                        "VsigGeschw": "-1"
                        if sig.vsig_geschw[i] == -1
                        else str(sig.vsig_geschw[i] / 3.6),
                    },
                )
            for i in range(0, len(sig.matrix)):
                for j in range(0, len(sig.matrix[i].spalten)):
                    if True:  # mz.vmax == 0 or mz.block or mz.gleis:
                        nstrecke_me = ET.SubElement(
                            nstrecke_signal,
                            "MatrixEintrag",
                            {
                                "MatrixGeschw": "-1"
                                if sig.matrix[i].spalten[j].vmax == -1
                                else str(sig.matrix[i].spalten[j].vmax / 3.6),
                                "Signalbild": str(sig.matrix[i].spalten[j].bild),
                            },
                        )
                        if sig.matrix[i].spalten[j].er1.nummer > 0:
                            nstrecke_me_er = ET.SubElement(
                                nstrecke_me,
                                "Ereignis",
                                {"Er": str(sig.matrix[i].spalten[j].er1.nummer)},
                            )
                            if sig.matrix[i].spalten[j].er1.wert > 0:
                                nstrecke_me_er.attrib["Wert"] = str(
                                    sig.matrix[i].spalten[j].er1.wert
                                )
                        if sig.matrix[i].spalten[j].er2.nummer > 0:
                            nstrecke_me_er = ET.SubElement(
                                nstrecke_me,
                                "Ereignis",
                                {"Er": str(sig.matrix[i].spalten[j].er2.nummer)},
                            )
                            if sig.matrix[i].spalten[j].er2.wert > 0:
                                nstrecke_me_er.attrib["Wert"] = str(
                                    sig.matrix[i].spalten[j].er2.wert
                                )

            if sig.master != 0:
                ET.SubElement(
                    ET.SubElement(
                        nstrecke_signal,
                        "KoppelSignal",
                        {
                            "ReferenzNr": str(
                                get_ref_nr(
                                    sig.master,
                                    RefTyp.SIGNAL
                                )
                            )
                        },
                    ),
                    "Datei",
                    {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                )
            # Referenzpunkt erzeugen

            allocate_refpunkt(nstrecke_strecke, elementnr, RefTyp.SIGNAL)

        register = int(f.readline())

        streckenelemente[elementnr]["aufloesepunkt"] = False

        if "NormEreignis" in streckenelemente[elementnr]:
            for i in range(len(streckenelemente[elementnr]["NormEreignis"])):
               if streckenelemente[elementnr]["NormEreignis"][i]["Er"] == 4:
                   streckenelemente[elementnr]["aufloesepunkt"] = True
                   allocate_refpunkt(nstrecke_strecke, elementnr, RefTyp.AUFLOESEPUNKT)

                   if register == 0:
                       # print(
                       #     f"kein Register an Auflöseelement {streckenelement['Nr']}, erfinde eins",
                       #     file=sys.stderr,
                       # )
                       register = regnr
                       regnr += 1

        if register != 0:
            nstrecke_norm.attrib["Reg"] = str(register)
            allocate_refpunkt(nstrecke_strecke, elementnr, RefTyp.REGISTER)

        streckenelemente[elementnr]["register"] = register

        # if signal["Vorhanden"] and Kombisignalvorhanden and len(signal["ls0"]) > 0:
        #     print(
        #         "In Streckenelement ",
        #         elementnr,
        #         " sind Signal und Kombisignal vorhanden und das Signal ist zweibegriffig",
        #     )

    # Streckenelementverbindung in Gegenrichtung ins Dictonary eintragen

    print("Streckenelementverbindung in Gegenrichtung ins Dictonary eintragen")
    for elem_nr, element in streckenelemente.items():
        for succ in element["NachNorm"]:
            if "NachGegen" in streckenelemente[succ]:
                preds = streckenelemente[succ]["NachGegen"]
                preds.append(elem_nr)
            else:
                streckenelemente[succ]["NachGegen"] = [elem_nr]

    # Hier kommt die Berechnung der Krümmung hin!!!!

    wb = Workbook()
    ws = wb.active
    ws.title = "Streckenelemente"
    ws.cell(row=1, column=1).value = "Streckenelementnummer"
    ws.cell(row=1, column=2).value = "x-Koordinate grün"
    ws.cell(row=1, column=3).value = "y-Koordinate grün"
    ws.cell(row=1, column=4).value = "z-Koordinate grün"
    ws.cell(row=1, column=5).value = "x-Koordinate blau"
    ws.cell(row=1, column=6).value = "y-Koordinate blau"
    ws.cell(row=1, column=7).value = "z-Koordinate blau"
    ws.cell(row=1, column=8).value = "Krümmung"
    ws.cell(row=1, column=9).value = "Nach Norm"
    ws.cell(row=1, column=10).value = "Nach Gegen"
    i = 1
    for elem_nr, element in streckenelemente.items():
        i += 1
        ws.cell(row=i, column=1).value = elem_nr
        if "gx" in element:
            ws.cell(row=i, column=2).value = element["gx"]
        if "gy" in element:
            ws.cell(row=i, column=3).value = element["gy"]
        if "gz" in element:
            ws.cell(row=i, column=4).value = element["gz"]
        if "bx" in element:
            ws.cell(row=i, column=5).value = element["bx"]
        if "by" in element:
            ws.cell(row=i, column=6).value = element["by"]
        if "bz" in element:
            ws.cell(row=i, column=7).value = element["bz"]
        if "kr" in element:
            ws.cell(row=i, column=8).value = element["kr"]
        if "NachNorm" in element:
            ws.cell(row=i, column=9).value = str(element["NachNorm"])
        if "NachGegen" in element:
            ws.cell(row=i, column=10).value = str(element["NachGegen"])
        # print("Element: ",elem_nr ,element)

    wb.save("strecke.xlsx")

    # Streckenelementverbindungen in XML-Datei eintragen

    print("Streckenelementverbindungen in XML-Datei eintragen")
    for elem_nr, element in streckenelemente.items():
        try:
            for i in element["NachNorm"]:
                ET.SubElement(nodes[elem_nr], "NachNorm").attrib["Nr"] = str(i)
        except KeyError:
            continue
        try:
            for i in element["NachGegen"]:
                ET.SubElement(nodes[elem_nr], "NachGegen").attrib["Nr"] = str(i)
        except KeyError:
            continue

    # Fahrstraßen

    def v_kleiner(v1, v2):
        if v2 == -1:
            return True
        elif v1 == -1:
            return False
        else:
            return v1 < v2

    def get_vsig_spalte(sig, v, ID):
        id_counter = 0
        for idx, vsig_geschw in enumerate(sig.vsig_geschw):
            if v == vsig_geschw:
                if id_counter == ID:
                    return idx
                else:
                    id_counter += 1

        spalte = 0
        spalte_geschw = -1
        for idx, vsig_geschw in enumerate(sig.vsig_geschw):
            if v != 0 and vsig_geschw != 0:
                if v_kleiner(vsig_geschw, v):
                    if vsig_geschw > spalte_geschw:
                        spalte = idx
                        spalte_geschw = vsig_geschw

        return spalte

    def get_aufloesepunkte_rek(elnr, startnr, nstrecke_fahrstrasse):
        while True:
            if elnr != startnr:
                if streckenelemente[elnr]["aufloesepunkt"]:
                    ET.SubElement(
                        ET.SubElement(
                            nstrecke_fahrstrasse,
                            "FahrstrAufloesung",
                            {"Ref": str(get_ref_nr(elnr, 5))},
                        ),
                        "Datei",
                        {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                    )
                    break

                if elnr in signale:
                    sig = signale[elnr]
                    if any(mz.vmax == 0 for mz in sig.matrix):
                        break

            succs = streckenelemente[elnr]["NachNorm"]
            if not succs:
                break
            for idx in range(1, len(succs)):
                get_aufloesepunkte_rek(startnr, succs[idx], nstrecke_fahrstrasse)
            elnr = succs[0]

    def get_fahrstr_rek(startnrs, elnr, nstrecke_fahrstrasse):
        while True:
            if elnr != startnrs[-1]:
                if streckenelemente[elnr]["register"]:
                    ET.SubElement(
                        ET.SubElement(
                            nstrecke_fahrstrasse,
                            "FahrstrRegister",
                            {"Ref": str(get_ref_nr(elnr, 2))},
                        ),
                        "Datei",
                        {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                    )

                if streckenelemente[elnr]["aufloesepunkt"]:
                    ET.SubElement(
                        ET.SubElement(
                            nstrecke_fahrstrasse,
                            "FahrstrTeilaufloesung",
                            {"Ref": str(get_ref_nr(elnr, 5))},
                        ),
                        "Datei",
                        {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                    )

                if elnr in fahrstrsignale:
                    ET.SubElement(
                        ET.SubElement(
                            nstrecke_fahrstrasse,
                            "FahrstrSignal",
                            {
                                "FahrstrSignalZeile": "1",
                                "Ref": str(
                                    get_ref_nr(elnr, RefTyp.SIGNAL_GEGENRICHTUNG)
                                ),
                            },
                        ),
                        "Datei",
                        {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                    )

                if elnr in signale:
                    sig = signale[elnr]
                    if True:  # if any(mz.vmax == 0 for mz in sig.matrix):
                        # TODO gibt es diese Unterscheidung auch in Zusi 2?
                        try:
                            startsig = signale[startnrs[-1]]
                        except KeyError:
                            startsig = None  # Aufgleispunkt

                        # Zielsignal verknüpfen
                        try:
                            zeile_v0 = next(
                                idx for idx, mz in enumerate(sig.matrix) if mz.vmax == 0
                            )
                        except StopIteration:
                            print(f"Signal ohne Zeile v=0", file=sys.stderr)
                            zeile_v0 = 0

                        ET.SubElement(
                            ET.SubElement(
                                nstrecke_fahrstrasse,
                                "FahrstrSignal",
                                {
                                    "FahrstrSignalZeile": str(zeile_v0),
                                    "Ref": str(get_ref_nr(elnr, 4)),
                                },
                            ),
                            "Datei",
                            {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                        )

                        # Startsignal und Vorsignale verknüpfen
                        if startsig is not None:
                            for idx, mz in enumerate(startsig.matrix):
                                if mz.block != sig.block or mz.gleis != sig.gleis:
                                    continue

                                ET.SubElement(
                                    ET.SubElement(
                                        nstrecke_fahrstrasse,
                                        "FahrstrSignal",
                                        {
                                            "FahrstrSignalZeile": str(idx),
                                            "Ref": str(get_ref_nr(startnrs[-1], 4)),
                                        },
                                    ),
                                    "Datei",
                                    {
                                        "Dateiname": z3streckendateiname_rel,
                                        "NurInfo": "1",
                                    },
                                )

                                # signalisierte Geschwindigkeit
                                hsig_geschw = None
                                ID = 0
                                for spalte, vsig_geschw in enumerate(
                                    startsig.vsig_geschw
                                ):
                                    if vsig_geschw == 0:
                                        hsig_geschw = (
                                            startsig.matrix[idx].spalten[spalte].vmax
                                        )
                                        ID = startsig.matrix[idx].spalten[spalte].id
                                        break
                                else:
                                    hsig_geschw = startsig.matrix[idx].spalten[0].vmax
                                    ID = startsig.matrix[idx].spalten[0].id

                                for vsig_nr in startsig.vsigs:
                                    try:
                                        vsig = signale[vsig_nr]
                                    except KeyError:
                                        try:
                                            vsig = anonymesignale[vsig_nr]
                                        except KeyError:
                                            # print(
                                            #     f"Kein Vorsignal an Element {vsig_nr}"
                                            # )
                                            continue

                                    ET.SubElement(
                                        ET.SubElement(
                                            nstrecke_fahrstrasse,
                                            "FahrstrVSignal",
                                            {
                                                "FahrstrSignalSpalte": str(
                                                    get_vsig_spalte(
                                                        vsig, hsig_geschw, ID
                                                    )
                                                ),
                                                "Ref": str(get_ref_nr(vsig_nr, 4)),
                                            },
                                        ),
                                        "Datei",
                                        {
                                            "Dateiname": z3streckendateiname_rel,
                                            "NurInfo": "1",
                                        },
                                    )
                                if hsig_geschw == 0:
                                    print(
                                        f" -> {sig.block} {sig.gleis}: vmax == 0 -> weiter",
                                        file=sys.stderr,
                                    )
                                    get_fahrstr_rek(
                                        startnrs + [elnr], elnr, nstrecke_fahrstrasse
                                    )
                                    return

                                break
                            else:
                                print(
                                    f"{startnrs[-1]}: keine zeile für Fahrweg nach {elnr} ({sig.block} {sig.gleis}) gefunden",
                                    file=sys.stderr,
                                )
                                return

                        ET.SubElement(
                            ET.SubElement(
                                nstrecke_fahrstrasse,
                                "FahrstrZiel",
                                {"Ref": str(get_ref_nr(elnr, 4))},
                            ),
                            "Datei",
                            {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                        )
                        fname = ""
                        for startnr in startnrs:
                            try:
                                startsig = signale[startnr]
                                fname += f"{startsig.block} {startsig.gleis} -> "
                            except KeyError:
                                fname += f"Aufgleispunkt -> "
                        fname += f"{sig.block} {sig.gleis}"
                        nstrecke_fahrstrasse.attrib["FahrstrName"] = fname

                        get_aufloesepunkte_rek(elnr, elnr, nstrecke_fahrstrasse)

                        nstrecke_fahrstrasse.attrib["FahrstrTyp"] = f"TypZug"
                        nstrecke_strecke.append(nstrecke_fahrstrasse)
                        # print(f" -> {fname}", file=sys.stderr)
                        break

            succs = streckenelemente[elnr]["NachNorm"]
            for idx, succ in enumerate(succs):
                if len(succs) == 1:
                    nstrecke_fahrstrasse2 = nstrecke_fahrstrasse
                else:
                    nstrecke_fahrstrasse2 = copy.deepcopy(nstrecke_fahrstrasse)

                succ_preds = streckenelemente[succ]["NachGegen"]
                if len(succ_preds) > 1:
                    ET.SubElement(
                        ET.SubElement(
                            nstrecke_fahrstrasse2,
                            "FahrstrWeiche",
                            {
                                "FahrstrWeichenlage": str(succ_preds.index(elnr) + 1),
                                "Ref": str(
                                    get_ref_nr(succ, RefTyp.WEICHE_GEGENRICHTUNG)
                                ),
                            },
                        ),
                        "Datei",
                        {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                    )

                if len(succs) == 1:
                    elnr = succs[0]
                    break
                else:
                    ET.SubElement(
                        ET.SubElement(
                            nstrecke_fahrstrasse2,
                            "FahrstrWeiche",
                            {
                                "FahrstrWeichenlage": str(idx + 1),
                                "Ref": str(get_ref_nr(elnr, RefTyp.WEICHE)),
                            },
                        ),
                        "Datei",
                        {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
                    )
                    get_fahrstr_rek(startnrs, succ, nstrecke_fahrstrasse2)
            else:
                break

    for elnr, sig in signale.items():
        if any(mz.vmax == 0 for mz in sig.matrix):
            # Hsig
            # print(f"{sig.block} {sig.gleis}", file=sys.stderr)
            nstrecke_fahrstrasse = ET.Element("Fahrstrasse")
            ET.SubElement(
                ET.SubElement(
                    nstrecke_fahrstrasse,
                    "FahrstrStart",
                    {"Ref": str(get_ref_nr(elnr, 4))},
                ),
                "Datei",
                {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
            )
            get_fahrstr_rek([elnr], elnr, nstrecke_fahrstrasse)

    for elnr in aufgleispunkte.values():
        # print(f"Aufgleispunkt {elnr}", file=sys.stderr)
        nstrecke_fahrstrasse = ET.Element("Fahrstrasse")
        ET.SubElement(
            ET.SubElement(
                nstrecke_fahrstrasse,
                "FahrstrStart",
                {"Ref": str(get_ref_nr(elnr, RefTyp.AUFGLEISPUNKT))},
            ),
            "Datei",
            {"Dateiname": z3streckendateiname_rel, "NurInfo": "1"},
        )
        get_fahrstr_rek([elnr], elnr, nstrecke_fahrstrasse)

    print(f"writing {z3streckendateiname_abs}", file=sys.stderr)
    os.makedirs(os.path.dirname(z3streckendateiname_abs), exist_ok=True)

    ET.indent(tree, space=" ", level=0)
    tree.write(z3streckendateiname_abs, encoding="UTF-8-SIG", xml_declaration=True)
    print(f"done", file=sys.stderr)
    return (z3streckendateiname_rel, rekursionstiefe)
